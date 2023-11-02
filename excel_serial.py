from openpyxl import load_workbook
from datetime import datetime

# TODO implement "class method" construction
# TODO add exceptions for each action, in case it fails
# TODO look into "active cell" in openpyxl to replace passing around row and column

FIRST_POSSIBLE_EMPTY_COLUMN = 2


class Item:
    def __init__(self, serial: str):
        self.serial = serial.strip()
        self._path = None
        self.wb = None
        self.sheet = None
        self.row = None
        self.column = None
        self.modelName = None
        self.deviceBirthday = None
        self.new = None
        self.prevName = None
        self.returned = None

        self._find_workbook()
        if not isinstance(self._wb, FileNotFoundError):
            self._find_sheet()
            if not isinstance(self.sheet, KeyError):
                self._find_row()
                self._assert_is_new()
                if self.new:
                    self.modelName = None
                    self.column = FIRST_POSSIBLE_EMPTY_COLUMN
                else:
                    self._fetch_device_info()
                    self._find_column()
                    self._fetch_prev_name()
                    self._assert_is_returned()

    def _create_path(self):
        path = "serial "
        self._path = path + self.serial[0:-3] + "000.xlsx"

    def _find_workbook(self):
        self._create_path()
        try:
            self._wb = load_workbook(self._path)
        except FileNotFoundError as e:
            self._wb = e

    def _find_sheet(self):
        targetSheet = self.serial[0:-2] + "00"
        try:
            self.sheet = self._wb[targetSheet]
        except KeyError as e:
            self.sheet = e

    def _find_row(self):
        cellVal = self.sheet.cell(row=1, column=1).value
        currentRow = 0
        while currentRow < 100 and cellVal != self.serial:
            currentRow += 1
            cellVal = self.sheet.cell(row=currentRow, column=1).value

        if cellVal != self.serial:
            self.row = -1
        else:
            self.row = currentRow

    def _assert_is_new(self):
        self.new = not self._not_last_cell(1)

    def _find_column(self):
        column = self._find_first_empty_cell(FIRST_POSSIBLE_EMPTY_COLUMN)
        result = self._not_last_cell(column)
        while result is not False:
            column = self._find_first_empty_cell(result)
            result = self._not_last_cell(column)

        self.column = column

    def _find_first_empty_cell(self, currentColumn):
        cellVal = self.sheet.cell(row=self.row, column=currentColumn).value
        while cellVal is not None:
            currentColumn += 1
            cellVal = self.sheet.cell(row=self.row, column=currentColumn).value
        return currentColumn

    def _not_last_cell(self, column):
        lastFound = None
        for i in range(1, 10):
            cellVal = self.sheet.cell(row=self.row, column=(column + i)).value
            if cellVal is not None:
                lastFound = i

        if lastFound is not None:
            return column + lastFound
        else:
            return False

    def _assert_is_returned(self):
        cellVal = self.sheet.cell(row=self.row, column=self.column - 1).value
        self.returned = cellVal == "הוחזר".strip()

    def _fetch_device_info(self):
        models = [
            "caneo",
            "domiflex",
            "exigo",
            "emineo",
            "cirrus",
            "marcus",
            "f3",
            "m1",
            "k300",
            "pt",
            "מדרגון",
            "eloflex",
            "adiflex",
        ]

        modelName = None
        currentColumn = 5
        cellVal1 = self.sheet.cell(row=self.row, column=currentColumn - 1).value
        cellVal2 = self.sheet.cell(row=self.row, column=currentColumn).value

        if type(cellVal1) is str:
            for x in models:
                if x in cellVal1.lower():
                    modelName = cellVal1
                    deviceBirthday = cellVal2
                    continue
        if modelName is None and type(cellVal2) is str:
            for x in models:
                if x in cellVal2.lower():
                    modelName = cellVal2
                    deviceBirthday = self.sheet.cell(
                        row=self.row, column=currentColumn + 1
                    ).value
                    continue
        if modelName is None:
            self.modelName = "לא נמצא"
        else:
            self.modelName = modelName.strip()

        try:
            self.deviceBirthday = deviceBirthday
        except UnboundLocalError:
            self.deviceBirthday = "לא נמצא"
        if isinstance(self.deviceBirthday, datetime):
            self.deviceBirthday = deviceBirthday.strftime("%d/%m/%y")

    def _fetch_prev_name(self):
        currentColumn = self.column
        for i in range(4, 8):
            if currentColumn - i >= 1:
                cellVal = self.sheet.cell(self.row, currentColumn - i).value
                if type(cellVal) is str and (
                    cellVal == "הוחזר".strip() or cellVal == self.serial.strip()
                ):
                    try:
                        self.prevName = self.sheet.cell(
                            self.row, currentColumn - i + 1
                        ).value.strip()
                    except AttributeError as e:
                        self.prevName = e
                        return
            else:
                continue
        self.prevName = "לא נמצא"

    def _check_unexpected_entry(self):
        value = self.sheet.cell(self.row, self.column).value
        return value is not None

    def _set_returned(self):
        if self._check_unexpected_entry():
            return False
        self.sheet.cell(self.row, self.column).value = "הוחזר"
        self._wb.save(self._path)
        return True

    def _update_info(self, name, id, date, model=None):
        if self._check_unexpected_entry():
            return False
        info = [name, id, model, date]
        for x in info:
            if x != "":
                self.sheet.cell(self.row, self.column).value = x
                self.column += 1
        self._wb.save(self._path)
        return True
