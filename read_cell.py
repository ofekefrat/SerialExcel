
import os, sys
from openpyxl import load_workbook

input = sys.argv[1]
path = "C:\\Users\\Rachel\\Desktop\\"
currentFileName = path +"test.xlsx"
wb = load_workbook(filename=currentFileName)
ws = wb['Sheet1']

print(ws['B2'].value)
sys.stdout.flush()