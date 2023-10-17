import os, sys
from openpyxl import load_workbook

path = "C:\\Users\\Rachel\\Desktop\\"
currentFileName = path +"test.xlsx"
wb = load_workbook(filename=currentFileName)
ws = wb['Sheet1']
ws['B2'] = 'Saved?'
wb.save(currentFileName)

print(ws['B2'].value)
