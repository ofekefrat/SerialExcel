from openpyxl import load_workbook
from flask import Flask, jsonify, render_template, request

def getSheet(serial):
    #only if 2400-03? 
    targetWorkbook = serial[0:-3] + "000"
    targetSheet = serial[0:-2] + "00"

    path = "C:\\Users\\Rachel\\Desktop\\"
    currentFileName = path + targetWorkbook
    workbook = load_workbook(filename=currentFileName)
    sheet = workbook[targetSheet]
    return sheet


def getRow(serial, sheet):
    currentRow=0
    while currentRow < 100 and cellVal != serial: 
        currentRow+=1
        cellVal = sheet.cell(row=currentRow, column=1).value

    if cellVal != serial:
        print("error: serial not found")
        return -1
    
    return currentRow


def isNew(sheet, row):
    for i in range(2, 5):
        cellVal = sheet.cell(row=row, column=i).value
        if cellVal is not None:
            return False
    return True


def getModelName(sheet, row):
    models = ["caneo", "domiflex", "exigo", "emineo", "cirrus", "marcus", "f3", "m1", "k300", "pt", "מדרגון", "eloflex"]

    currentColumn=4
    cellVal1 = sheet.cell(row=row, column=currentColumn-1).value
    cellVal2 = sheet.cell(row=row, column=currentColumn).value
    for i in models:
        if models[i] in cellVal1.lower():
            modelName = cellVal1
            continue
        elif models[i] in cellVal2.lower():
            modelName = cellVal2
            continue
    
def getColumn(sheet, row):
    currentColumn=2
    while cellVal is not None:
        currentColumn+=1
        cellVal = sheet.cell(row=row, column=currentColumn).value
    
    cellVal = sheet.cell(row=row, column=currentColumn-1).value
    if cellVal != "הוחזר":
        print("error: device not returned")
        modelName = None

    return modelName
    
# MAIN
app = Flask(__name__)

@app.route('/find_serial')
def find_serial():
    serial = request.args.get('serial', 0, type=str)
    # sheet = getSheet(serial)
    # row = getRow(serial, sheet)
    # if isNew(sheet, row):
    #     modelName = None
    #     column = 2
    # else:
    #     modelName = getModelName(sheet, row)
    #     column = getColumn(sheet, row)
        
    # data = {
    #     "row": row,
    #     "column": column,
    #     "modelName": modelName
    # }
    data = {
        "modelName": serial + "THANKS",
        "row": 1
    }
    return jsonify(data)

@app.route('/')
def index():
    return render_template('injector.html')
