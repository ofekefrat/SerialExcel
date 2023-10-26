from openpyxl import load_workbook
from datetime import datetime

#TODO add exceptions for each action, in case it fails

class Item:
    def __init__(self, serial: str):
        self.serial = serial.strip()
        self.path = None
        self.wb = None
        self.sheet = None
        self.row = None
        self.column = None
        self.modelName = None
        self.deviceBirthday = None
        self.new = None
        self.prevName = None
        self.returned = None

        self.findWorkbook()
        if not isinstance(self.wb, FileNotFoundError):
            self.findSheet()
            if not isinstance(self.sheet, KeyError):
                self.findRow()
                self.assertIsNew()
                if self.new:
                    self.modelName = None
                    self.column = 2
                else:
                    self.fetchDeviceInfo()
                    self.findColumn()
                    self.fetchPrevName()
                    self.assertIsReturned()
    

    def createPath(self):
        path = "serial "
        self.path = path + self.serial[0:-3] + "000.xlsx"
        

    def findWorkbook(self):
        self.createPath()
        try:
            self.wb = load_workbook(self.path)
        except FileNotFoundError as e:
            self.wb = e
            

    def findSheet(self):
        targetSheet = self.serial[0:-2] + "00"
        try:
            self.sheet = self.wb[targetSheet]   
        except KeyError as e:
            self.sheet = e


    def findRow(self):
        cellVal = self.sheet.cell(row=1, column=1).value
        currentRow=0
        while currentRow < 100 and cellVal != self.serial: 
            currentRow+=1
            cellVal = self.sheet.cell(row=currentRow, column=1).value

        if cellVal != self.serial:
            self.row = -1
        else:
            self.row = currentRow


    def assertIsNew (self):
        self.new = not self.not_last_cell(1)


    def findColumn(self):
        column = self.find_first_empty_cell()
        result = self.not_last_cell(column)
        while result is not False:
            column = self.find_first_empty_cell()
            result = self.not_last_cell(column)

        self.column = column

    
    def assertIsReturned(self):
        cellVal = self.sheet.cell(row=self.row, column=self.column-1).value
        self.returned = (cellVal == "הוחזר".strip())
        

    def fetchDeviceInfo(self):
        models = ["caneo", "domiflex", "exigo", "emineo", "cirrus", "marcus", "f3", "m1", "k300", "pt", "מדרגון", "eloflex", "adiflex"]

        modelName = None
        currentColumn=5
        cellVal1 = self.sheet.cell(row=self.row, column=currentColumn-1).value
        cellVal2 = self.sheet.cell(row=self.row, column=currentColumn).value

        if type(cellVal1) is str:
            for x in models:
                if x in cellVal1.lower():
                    modelName = cellVal1
                    deviceBirthday = cellVal2
                    continue
        if modelName is None and type(cellVal2) is str:
            for x in cellVal2.lower():
                if x in cellVal2.lower():
                    modelName = cellVal2
                    deviceBirthday = self.sheet.cell(row=self.row, column=currentColumn+1).value
                    continue
        if modelName is None:
            self.modelName = "לא נמצא"
        else:
            self.modelName = modelName

        self.deviceBirthday = deviceBirthday
        if isinstance(deviceBirthday, datetime):
            self.deviceBirthday = deviceBirthday.strftime("%d/%m/%y")
        

    def fetchPrevName(self):
        currentColumn = self.column
        for i in range(4, 8):
            cellVal = self.sheet.cell(self.row, currentColumn-i).value
            if type(cellVal) is str and (cellVal == "הוחזר".strip() or cellVal == self.serial.strip()):
                self.prevName = self.sheet.cell(self.row, currentColumn-i+1).value
                return
        self.prevName = "לא נמצא"
                

    def find_first_empty_cell(self):
        currentColumn=2
        cellVal = self.sheet.cell(row=self.row, column=currentColumn).value
        while cellVal is not None:
            currentColumn+=1
            cellVal = self.sheet.cell(row=self.row, column=currentColumn).value
        return currentColumn


    def not_last_cell(self, column):
        for i in range(1, 10):
            cellVal = self.sheet.cell(row=self.row, column=(column+i)).value
            if cellVal is not None:
                return column+i
            else:
                return False


    def checkUnexpectedEntry(self):
        value = self.sheet.cell(self.row, self.column).value
        return value is not None


    def setReturned(self):
        if self.checkUnexpectedEntry():
            return False
        self.sheet.cell(self.row, self.column).value = "הוחזר"
        self.wb.save(self.path)
        return True


    def updateInfo(self, name, id, date, model=None):
        if self.checkUnexpectedEntry():
            return False
        info = [name, id, model, date]
        for x in info:
            if x != "":
                self.sheet.cell(self.row, self.column).value = x
                self.column+=1
        self.wb.save(self.path)
        return True